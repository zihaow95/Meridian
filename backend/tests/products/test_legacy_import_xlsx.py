"""Excel/XLSX import template and multipart evidence."""

from __future__ import annotations

import io
from collections.abc import Callable

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.platform.application.command import CommandContext
from apps.products.models import ImportBatchStatus
from apps.products.services.import_batch import CreateProductImportBatch
from apps.products.services.import_template import (
    ALL_COLUMNS,
    build_import_template_xlsx,
    parse_import_csv,
    parse_import_xlsx,
    sample_import_csv,
)


@pytest.fixture
def migration_operator(product_director, grant_action: Callable[..., None]):
    grant_action(product_director, "migration.upload", "migration", role_code="PRODUCT_DIRECTOR")
    grant_action(product_director, "migration.review", "migration", role_code="PRODUCT_DIRECTOR")
    return product_director


@pytest.mark.django_db
def test_xlsx_template_round_trip_matches_csv_rows() -> None:
    template_bytes = build_import_template_xlsx()
    workbook = load_workbook(filename=io.BytesIO(template_bytes), read_only=True)
    sheet = workbook.active
    assert sheet is not None
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert headers == list(ALL_COLUMNS)

    xlsx_rows = parse_import_xlsx(content=template_bytes)
    csv_rows = parse_import_csv(content=sample_import_csv())
    assert len(xlsx_rows) == 1
    assert xlsx_rows[0].normalized_payload["name"] == "Legacy yogurt"
    assert (
        xlsx_rows[0].normalized_payload["category_code"]
        == csv_rows[0].normalized_payload["category_code"]
    )


@pytest.mark.django_db
def test_create_import_batch_from_xlsx_content(migration_operator) -> None:
    batch = CreateProductImportBatch(
        context=CommandContext.for_actor(migration_operator),
        xlsx_content=build_import_template_xlsx(),
        source_filename="legacy.xlsx",
    ).execute()
    assert batch.status == ImportBatchStatus.PARSED
    assert batch.total_count == 1
    assert batch.success_count == 1
    assert batch.source_filename == "legacy.xlsx"


@pytest.mark.django_db
def test_multipart_xlsx_upload_and_template_download(
    api_client: APIClient,
    migration_operator,
) -> None:
    api_client.force_authenticate(user=migration_operator)

    template = api_client.get("/api/v1/product-import-template")
    assert template.status_code == 200
    assert (
        template["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=io.BytesIO(template.content), read_only=True)
    assert workbook.active is not None

    upload = api_client.post(
        "/api/v1/product-import-batches",
        {
            "file": SimpleUploadedFile(
                "legacy.xlsx",
                build_import_template_xlsx(),
                content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            )
        },
        format="multipart",
    )
    assert upload.status_code == 201
    body = upload.json()
    assert body["status"] == ImportBatchStatus.PARSED
    assert body["total_count"] == 1
    assert body["success_count"] == 1
