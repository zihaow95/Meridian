"""Legacy product import CSV template definition and parsing."""

from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass
from typing import Any

IMPORT_TEMPLATE_VERSION = "legacy-product-v1"

REQUIRED_COLUMNS = ("name", "category_code")
OPTIONAL_COLUMNS = (
    "business_no",
    "external_id",
    "brand_code",
    "sku_code",
    "barcode",
    "specification",
    "net_content_value",
    "net_content_unit",
)
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


@dataclass(frozen=True)
class ParsedImportRow:
    row_number: int
    raw_row_digest: str
    normalized_payload: dict[str, Any]
    validation_errors: list[dict[str, str]]


def digest_row(*, row_number: int, values: dict[str, str]) -> str:
    payload = f"{row_number}:{sorted(values.items())}".encode()
    return hashlib.sha256(payload).hexdigest()


def parse_import_csv(*, content: str) -> list[ParsedImportRow]:
    reader = csv.DictReader(io.StringIO(content))
    if reader.fieldnames is None:
        return []
    missing = [column for column in REQUIRED_COLUMNS if column not in reader.fieldnames]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: list[ParsedImportRow] = []
    for index, raw in enumerate(reader, start=1):
        values = {
            column: (raw.get(column) or "").strip() for column in ALL_COLUMNS if column in raw
        }
        errors: list[dict[str, str]] = []
        if not values.get("name"):
            errors.append({"field": "name", "message": "Name is required."})
        if not values.get("category_code"):
            errors.append({"field": "category_code", "message": "Category code is required."})
        rows.append(
            ParsedImportRow(
                row_number=index,
                raw_row_digest=digest_row(row_number=index, values=values),
                normalized_payload=values,
                validation_errors=errors,
            )
        )
    return rows


def sample_import_csv() -> str:
    return (
        "name,category_code,business_no,brand_code,sku_code,barcode,specification\n"
        "Legacy yogurt,YOGURT,LEG-001,BRAND-A,SKU-LEG-001,6900000000001,120g\n"
        "Legacy milk,DAIRY,LEG-002,BRAND-B,SKU-LEG-002,6900000000002,250ml\n"
    )


def build_import_template_xlsx() -> bytes:
    """Build the standard Excel import template for download."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "products"
    sheet.append(list(ALL_COLUMNS))
    sheet.append(
        [
            "Legacy yogurt",
            "YOGURT",
            "LEG-001",
            "",
            "BRAND-A",
            "SKU-LEG-001",
            "6900000000001",
            "120g",
            "",
            "",
        ]
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_import_xlsx(*, content: bytes) -> list[ParsedImportRow]:
    """Parse an Excel workbook into the same row shape as CSV imports."""
    from openpyxl import load_workbook

    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(cell or "").strip() for cell in header_row]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: list[ParsedImportRow] = []
    for index, raw in enumerate(rows_iter, start=1):
        mapped = {
            headers[col_index]: ("" if value is None else str(value).strip())
            for col_index, value in enumerate(raw)
            if col_index < len(headers)
        }
        values = {column: mapped.get(column, "") for column in ALL_COLUMNS}
        errors: list[dict[str, str]] = []
        if not values.get("name"):
            errors.append({"field": "name", "message": "Name is required."})
        if not values.get("category_code"):
            errors.append({"field": "category_code", "message": "Category code is required."})
        rows.append(
            ParsedImportRow(
                row_number=index,
                raw_row_digest=digest_row(row_number=index, values=values),
                normalized_payload=values,
                validation_errors=errors,
            )
        )
    return rows
