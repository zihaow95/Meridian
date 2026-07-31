"""Parse controlled CSV/Excel files into row dicts for ingestion."""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook

from apps.documents.models import DocumentVersion, StorageStatus, VersionStatus
from apps.documents.storage.factory import get_file_storage
from apps.platform.api.errors import ValidationFailedError


def assert_controlled_active_document_version(version: DocumentVersion) -> None:
    if version.status != VersionStatus.CONTROLLED:
        raise ValidationFailedError(message="Document version must be CONTROLLED.")
    if version.file_object.storage_status != StorageStatus.ACTIVE:
        raise ValidationFailedError(message="Document file must be ACTIVE.")


def parse_rows_from_document_version(version: DocumentVersion) -> list[dict[str, Any]]:
    assert_controlled_active_document_version(version)
    path = get_file_storage().final_path_for(version.file_object.object_key)
    if not path.exists():
        raise ValidationFailedError(message="Ingestion input file is missing.")
    name = (version.original_filename or path.name).lower()
    mime = (version.detected_mime_type or version.declared_mime_type or "").lower()
    raw = path.read_bytes()
    if name.endswith(".csv") or mime in {"text/csv", "application/csv"}:
        return _parse_csv_bytes(raw)
    if name.endswith(".xlsx") or "spreadsheetml" in mime or name.endswith(".xls"):
        return _parse_xlsx_bytes(raw)
    raise ValidationFailedError(message="Unsupported ingestion file type.")


def _parse_csv_bytes(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [{str(k): ("" if v is None else str(v)) for k, v in row.items()} for row in reader]


def _parse_xlsx_bytes(raw: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValidationFailedError(message="Excel workbook has no active sheet.")
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration as exc:
        raise ValidationFailedError(message="Excel file has no header row.") from exc
    parsed: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(cell is None or str(cell).strip() == "" for cell in values):
            continue
        row: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else None
            row[header] = "" if value is None else str(value)
        parsed.append(row)
    return parsed
